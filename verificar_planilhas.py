from openpyxl import load_workbook

# Carrega o arquivo Excel
arquivo = "C:\\Users\\lz6321\\OneDrive - grendene.com.br\\Documentos\\Automações\\Bens.xlsx"
wb = load_workbook(arquivo)

# Mostra os nomes de todas as planilhas
print("Planilhas disponíveis:")
for planilha in wb.sheetnames:
    print(f"- {planilha}")
